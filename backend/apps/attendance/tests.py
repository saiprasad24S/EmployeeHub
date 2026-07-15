from datetime import date, datetime, timedelta
from io import BytesIO

from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from django.utils import timezone
from openpyxl import load_workbook
from PIL import Image
from rest_framework.exceptions import ValidationError

from apps.accounts.models import Employee
from apps.assignments.models import Assignment
from apps.attendance.models import Session
from apps.attendance.services import (
    annotate_image,
    generate_attendance_export,
    get_employee_presence_summary,
    validate_geofence,
)


class GeofenceTests(TestCase):
    def setUp(self):
        self.employee = Employee.objects.create(
            employee_id="EMP001",
            name="Test Employee",
            email="employee@example.com",
            department="Field",
            designation="Nurse",
        )
        self.assignment = Assignment.objects.create(
            employee=self.employee,
            patient_name="Patient One",
            patient_address="Test Address",
            latitude=12.9716,
            longitude=77.5946,
            radius=150,
            visit_date=timezone.localdate(),
            status=Assignment.Status.ACTIVE,
        )

    def test_validate_geofence_allows_nearby_location(self):
        validate_geofence(self.assignment, 12.9720, 77.5950)

    def test_validate_geofence_rejects_far_location(self):
        with self.assertRaises(ValidationError):
            validate_geofence(self.assignment, 13.1000, 77.8000)

    def test_annotate_image_adds_timestamp_and_location_overlay(self):
        image = Image.new('RGB', (300, 200), color='blue')
        buffer = BytesIO()
        image.save(buffer, format='JPEG')
        uploaded = SimpleUploadedFile('photo.jpg', buffer.getvalue(), content_type='image/jpeg')

        annotated = annotate_image(uploaded, timestamp='2024-01-01 10:00', location='12.9716, 77.5946')

        self.assertIsNotNone(annotated)
        output = Image.open(annotated)
        self.assertEqual(output.size[0], 300)
        self.assertEqual(output.size[1], 200)

    def test_generate_attendance_export_creates_workbook_rows(self):
        employee = Employee.objects.create(
            employee_id='EMP002',
            name='Jane Doe',
            email='jane@example.com',
            phone='9876543210',
        )
        session = Session.objects.create(employee=employee, login_time=datetime(2024, 1, 2, 8, 0), logout_time=datetime(2024, 1, 2, 16, 0), is_active=False)
        session.created_at = datetime(2024, 1, 2, 8, 0)
        session.save(update_fields=['created_at'])

        workbook_bytes = generate_attendance_export(date(2024, 1, 1), date(2024, 1, 3))
        workbook = load_workbook(BytesIO(workbook_bytes))
        sheet = workbook.active
        self.assertGreater(sheet.max_row, 1)
        self.assertIn(sheet.cell(row=2, column=3).value, {'EMP001', 'EMP002'})

    def test_get_employee_presence_summary_keeps_active_session_present_across_days(self):
        employee = Employee.objects.create(
            employee_id='EMP003',
            name='Night Shift',
            email='night@example.com',
            phone='9988776655',
        )
        session = Session.objects.create(
            employee=employee,
            is_active=True,
        )
        session.login_time = timezone.make_aware(datetime(2024, 1, 2, 22, 30))
        session.save(update_fields=['login_time'])

        summary = get_employee_presence_summary(employee)

        self.assertTrue(summary['is_present'])
        self.assertEqual(summary['status'], 'Present')
        self.assertEqual(summary['check_in_time'], timezone.make_aware(datetime(2024, 1, 2, 22, 30)))
