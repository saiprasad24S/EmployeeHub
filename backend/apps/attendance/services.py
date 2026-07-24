from __future__ import annotations

from datetime import date, datetime, timedelta
from io import BytesIO
from typing import Any

import requests
from django.conf import settings
from django.db import transaction
from django.db.models import Q
from django.utils import timezone
from openpyxl import Workbook
from rest_framework.exceptions import ValidationError

from apps.accounts.models import Employee
from apps.assignments.models import Assignment
from apps.attendance.models import Attendance, Session
from apps.common.cloudinary_service import upload_attendance_image, upload_profile_image
from apps.common.utils import distance_meters
from apps.tracking.models import LocationLog
from apps.vision.services import FaceService, LivenessService


face_service = FaceService()
liveness_service = LivenessService()


def get_active_assignment(employee: Employee) -> Assignment | None:
    today = timezone.localdate()
    return (
        Assignment.objects.filter(employee=employee, visit_date=today)
        .exclude(status=Assignment.Status.CANCELLED)
        .order_by("-created_at")
        .first()
    )


def geocode_address(address: str) -> tuple[float, float] | None:
    if not address:
        return None
    try:
        response = requests.get(
            "https://nominatim.openstreetmap.org/search",
            params={"q": address, "format": "jsonv2", "limit": 1},
            headers={"User-Agent": "EmployeeHub/1.0"},
            timeout=10,
        )
        response.raise_for_status()
        data = response.json()
        if not data:
            return None
        return float(data[0]["lat"]), float(data[0]["lon"])
    except Exception:
        return None


def ensure_default_coordinates(employee: Employee) -> bool:
    if employee.default_latitude is not None and employee.default_longitude is not None:
        return True
    if not employee.default_address:
        return False
    coords = geocode_address(employee.default_address)
    if not coords:
        return False
    employee.default_latitude, employee.default_longitude = coords
    employee.save(update_fields=["default_latitude", "default_longitude"])
    return True


def validate_geofence(assignment: Assignment, latitude: float, longitude: float, accuracy: float | None = None) -> None:
    radius = assignment.radius or settings.DEFAULT_GEOFENCE_RADIUS_METERS
    distance = distance_meters(float(latitude), float(longitude), float(assignment.latitude), float(assignment.longitude))
    buffer = max(10.0, (accuracy or 0) * 1.5)
    if distance > radius + buffer:
        raise ValidationError({"detail": f"Geofence Verification Failed: You are outside the patient's scheduled range by {round(distance - radius, 2)} meters. Attendance was not marked."})


def upload_selfie(
    image_file,
    folder: str,
    *,
    employee: Employee | None = None,
    timestamp: str | None = None,
    location: str | None = None,
) -> dict[str, str]:
    from apps.accounts.models import Employee

    if not employee and hasattr(image_file, "employee") and isinstance(image_file.employee, Employee):
        employee = image_file.employee

    employee_id = employee.employee_id if employee else "unknown"
    employee_name = employee.name if employee else "Unknown"

    attendance_type = "checkin" if folder == "attendance" else "checkout"
    result = upload_attendance_image(
        image_file,
        employee_id=employee_id,
        attendance_type=attendance_type,
        timestamp=timestamp,
        address=location,
        employee_name=employee_name,
    )
    return {"url": result["url"], "public_id": result["public_id"]}


def upload_profile_photo(image_file, *, employee_id: str, employee_name: str | None = None) -> dict[str, str]:
    result = upload_profile_image(image_file, employee_id=employee_id, employee_name=employee_name)
    return {"url": result["url"], "public_id": result["public_id"]}


def get_employee_presence_summary(employee: Employee, *, reference_time: datetime | None = None) -> dict[str, Any]:
    reference_time = reference_time or timezone.now()
    session = Session.objects.filter(employee=employee, is_active=True).order_by('-login_time').first()
    if not session:
        session = Session.objects.filter(employee=employee).order_by('-login_time').first()
    if not session:
        return {
            'is_present': False,
            'status': 'Absent',
            'check_in_time': None,
            'check_out_time': None,
            'session_duration_seconds': 0,
            'session': None,
        }

    if session.is_active:
        duration_seconds = max(int((reference_time - session.login_time).total_seconds()), 0)
        return {
            'is_present': True,
            'status': 'Present',
            'check_in_time': session.login_time,
            'check_out_time': None,
            'session_duration_seconds': duration_seconds,
            'session': session,
        }

    logout_time = session.logout_time or session.login_time
    duration_seconds = max(int((logout_time - session.login_time).total_seconds()), 0)
    return {
        'is_present': False,
        'status': 'Absent',
        'check_in_time': session.login_time,
        'check_out_time': logout_time,
        'session_duration_seconds': duration_seconds,
        'session': session,
    }


def get_session_for_date(employee: Employee, target_date: date) -> Session | None:
    return (
        Session.objects.filter(employee=employee, login_time__date__lte=target_date)
        .filter(Q(logout_time__isnull=True) | Q(logout_time__date__gte=target_date))
        .order_by('-login_time')
        .first()
    )


def generate_attendance_export(start_date: date, end_date: date) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    # Report Header Block
    ws.cell(row=1, column=1, value="Skandan Home Carre Clinic LLP").font = Font(name="Calibri", size=16, bold=True, color="1F497D")
    ws.cell(row=2, column=1, value="Attendance Report").font = Font(name="Calibri", size=13, bold=True, color="595959")
    ws.cell(row=3, column=1, value=f"Date Range: {start_date.strftime('%d-%b-%Y')} to {end_date.strftime('%d-%b-%Y')}").font = Font(name="Calibri", size=11, italic=True)
    ws.cell(row=4, column=1, value=f"Generated On: {timezone.localtime(timezone.now()).strftime('%d-%b-%Y %I:%M %p')} | Generated By: HR Admin").font = Font(name="Calibri", size=10, color="7F7F7F")

    # Generate Date list
    date_list = []
    curr = start_date
    while curr <= end_date:
        date_list.append(curr)
        curr += timedelta(days=1)

    # Styling definitions
    header_fill = PatternFill(start_color="6B2FA0", end_color="6B2FA0", fill_type="solid")
    sub_header_fill = PatternFill(start_color="F2EBF9", end_color="F2EBF9", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    sub_header_font = Font(name="Calibri", size=10, bold=True, color="6B2FA0")
    bold_font = Font(name="Calibri", size=11, bold=True)
    regular_font = Font(name="Calibri", size=10)
    
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9")
    )
    
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center")

    # Row 6: Main Header Row
    ws.cell(row=6, column=1, value="Employee ID")
    ws.cell(row=6, column=2, value="Employee Name")

    # Merge Row 6 & Row 7 for first two columns
    ws.merge_cells(start_row=6, start_column=1, end_row=7, end_column=1)
    ws.merge_cells(start_row=6, start_column=2, end_row=7, end_column=2)

    col_idx = 3
    for dt in date_list:
        date_str = dt.strftime("%d-%b-%Y")
        ws.cell(row=6, column=col_idx, value=date_str)
        ws.merge_cells(start_row=6, start_column=col_idx, end_row=6, end_column=col_idx + 2)
        
        # Row 7: Sub-headers
        ws.cell(row=7, column=col_idx, value="Check In")
        ws.cell(row=7, column=col_idx + 1, value="Check Out")
        ws.cell(row=7, column=col_idx + 2, value="Working Hours")
        
        col_idx += 3

    # Apply Header Styles
    for r in range(6, 8):
        for c in range(1, col_idx):
            cell = ws.cell(row=r, column=c)
            cell.alignment = center_align
            cell.border = thin_border
            if r == 6:
                cell.fill = header_fill
                cell.font = header_font
            else:
                if c > 2:
                    cell.fill = sub_header_fill
                    cell.font = sub_header_font

    # Populate Data Rows
    employees = Employee.objects.all().order_by("employee_id")
    row_idx = 8

    for emp in employees:
        ws.cell(row=row_idx, column=1, value=emp.employee_id).alignment = left_align
        ws.cell(row=row_idx, column=2, value=emp.name).alignment = left_align

        ws.cell(row=row_idx, column=1).font = bold_font
        ws.cell(row=row_idx, column=2).font = bold_font

        c_idx = 3
        for dt in date_list:
            session = Session.objects.filter(employee=emp, login_time__date=dt).first()
            if not session:
                session = get_session_for_date(emp, dt)

            if not session:
                check_in_val = "-"
                check_out_val = "-"
                hours_val = "Absent"
            else:
                login_time = session.login_time
                logout_time = session.logout_time
                
                check_in_val = timezone.localtime(login_time).strftime("%I:%M %p") if login_time else "-"
                
                if session.is_active or not logout_time:
                    check_out_val = "Active"
                    diff = timezone.now() - login_time
                    h = int(diff.total_seconds() // 3600)
                    m = int((diff.total_seconds() % 3600) // 60)
                    hours_val = f"{h}h {m}m"
                else:
                    check_out_val = timezone.localtime(logout_time).strftime("%I:%M %p")
                    diff = logout_time - login_time
                    h = int(diff.total_seconds() // 3600)
                    m = int((diff.total_seconds() % 3600) // 60)
                    hours_val = f"{h}h {m}m"

            c1 = ws.cell(row=row_idx, column=c_idx, value=check_in_val)
            c2 = ws.cell(row=row_idx, column=c_idx + 1, value=check_out_val)
            c3 = ws.cell(row=row_idx, column=c_idx + 2, value=hours_val)

            for c in (c1, c2, c3):
                c.alignment = center_align
                c.font = regular_font
                c.border = thin_border
                if hours_val == "Absent":
                    c.font = Font(name="Calibri", size=10, color="9C0006")

            c_idx += 3

        row_idx += 1

    # Apply borders to employee columns in data region
    for r in range(8, row_idx):
        ws.cell(row=r, column=1).border = thin_border
        ws.cell(row=r, column=2).border = thin_border

    # Freeze Panes below headers and after column B
    ws.freeze_panes = "C8"

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row < 6:
                continue
            val = str(cell.value or '')
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


@transaction.atomic
def start_session(employee: Employee) -> Session:
    active_session = Session.objects.select_for_update().filter(employee=employee, is_active=True).first()
    if active_session:
        return active_session
    return Session.objects.create(employee=employee, is_active=True)


@transaction.atomic
def end_session(employee: Employee) -> Session:
    session = Session.objects.select_for_update().filter(employee=employee, is_active=True).first()
    if not session:
        raise ValidationError({"detail": "No active session found."})
    session.is_active = False
    session.logout_time = timezone.now()
    session.save(update_fields=["is_active", "logout_time"])
    return session


def record_attendance(
    *,
    employee: Employee,
    assignment: Assignment | None,
    session: Session | None,
    attendance_type: str,
    photo_url: str,
    photo_public_id: str = "",
    latitude: float,
    longitude: float,
    address: str,
    status: str,
    remarks: str = "",
) -> Attendance:
    return Attendance.objects.create(
        employee=employee,
        assignment=assignment,
        session=session,
        attendance_type=attendance_type,
        photo_url=photo_url,
        photo_public_id=photo_public_id,
        latitude=latitude,
        longitude=longitude,
        address=address,
        status=status,
        remarks=remarks,
    )


def annotate_image(image_file, *, timestamp: str | None = None, location: str | None = None):
    from PIL import Image as PILImage, ImageDraw, ImageFont

    image = PILImage.open(image_file).convert("RGBA")
    draw = ImageDraw.Draw(image)
    font = ImageFont.load_default()

    text = []
    if timestamp:
        text.append(timestamp)
    if location:
        text.append(location)
    if text:
        draw.text((10, 10), "\n".join(text), fill=(255, 255, 255, 255), font=font)
    output = BytesIO()
    image.convert("RGB").save(output, format="JPEG")
    output.seek(0)
    return output


def validate_liveness(image_file, liveness_score: float | None = None) -> None:
    if not liveness_service.is_live(image_file=image_file, liveness_score=liveness_score):
        raise ValidationError({"detail": "Liveness check failed."})


def log_location(
    *,
    session: Session,
    employee: Employee,
    latitude: float,
    longitude: float,
    accuracy: float | None = None,
    speed: float | None = None,
    battery_percentage: int | None = None,
    is_mock: bool = False,
    client_timestamp: datetime | None = None,
) -> LocationLog:
    if is_mock:
        raise ValidationError({"detail": "Mock location detected."})
    if client_timestamp:
        skew = abs((timezone.now() - client_timestamp).total_seconds())
        if skew > 300:
            raise ValidationError({"detail": "Timestamp skew is too large."})
    return LocationLog.objects.create(
        session=session,
        employee=employee,
        latitude=latitude,
        longitude=longitude,
        accuracy=accuracy,
        speed=speed,
        battery_percentage=battery_percentage,
    )
